"""
update_managers.py — HMO Equipment Management System
=====================================================
Đồng bộ cột "CB quản lý hiện tại" từ HMO_Master_Equipment_Database.xlsx
vào field "manager" trong JSON của QR_Landing_Page.html.

Cách dùng:
    python update_managers.py

Yêu cầu:
    pip install openpyxl

Script tự động:
    1. Đọc sheet Master_Data trong file .xlsx
    2. Trích xuất mapping: Mã QR → CB quản lý hiện tại
    3. Tìm và cập nhật field "manager" trong EQUIPMENT JSON của landing page
    4. Tạo file backup trước khi ghi đè
    5. In báo cáo: bao nhiêu thiết bị được cập nhật, bao nhiêu không khớp
"""

import re
import json
import shutil
import openpyxl
from pathlib import Path
from datetime import datetime

# ── Cấu hình đường dẫn ──────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
XLSX_FILE   = BASE_DIR / "HMO_Master_Equipment_Database.xlsx"
HTML_FILE   = BASE_DIR / "QR_Landing_Page.html"
BACKUP_DIR  = BASE_DIR / "backups"

SHEET_NAME  = "Master_Data"
COL_QR      = "Mã QR"
COL_MANAGER = "CB quản lý hiện tại"

# ── Regex tìm EQUIPMENT JSON block trong HTML ────────────────────────────────
# Khớp: const EQUIPMENT = { ... };
EQUIPMENT_RE = re.compile(
    r'(const EQUIPMENT\s*=\s*)(\{.*?\})\s*;',
    re.DOTALL
)

# ── Regex cập nhật field "manager" theo từng QR code ────────────────────────
# Khớp: "HMO-XXX-YYYY": { ... "manager": "...", ... }
# Dùng để patch từng entry mà không cần parse toàn bộ JSON (an toàn hơn)
def make_manager_re(qr_code: str) -> re.Pattern:
    escaped = re.escape(qr_code)
    return re.compile(
        r'("' + escaped + r'"\s*:\s*\{[^}]*?"manager"\s*:\s*")([^"]*?)(")',
        re.DOTALL
    )


def read_managers_from_xlsx(xlsx_path: Path) -> dict[str, str]:
    """Đọc sheet Master_Data, trả về dict {qr_code: manager_name}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}' trong file xlsx.")
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    try:
        qr_idx  = headers.index(COL_QR)
        mgr_idx = headers.index(COL_MANAGER)
    except ValueError as e:
        raise ValueError(f"Không tìm thấy cột: {e}") from e

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        qr  = row[qr_idx]
        mgr = row[mgr_idx]
        if qr:  # bỏ qua dòng trống
            mgr_value = str(mgr).strip() if mgr else "Chưa phân công"
            mapping[str(qr).strip()] = mgr_value
    wb.close()
    return mapping


def backup_html(html_path: Path) -> Path:
    """Tạo bản backup trước khi chỉnh sửa."""
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"QR_Landing_Page_{ts}.html"
    shutil.copy2(html_path, backup_path)
    return backup_path


def update_managers_in_html(html_content: str, mapping: dict[str, str]) -> tuple[str, list, list]:
    """
    Cập nhật field manager trong HTML.
    Trả về: (html_mới, danh_sách_đã_cập_nhật, danh_sách_không_khớp)
    """
    updated  = []
    skipped  = []

    for qr_code, new_manager in mapping.items():
        pattern = make_manager_re(qr_code)
        match   = pattern.search(html_content)

        if match:
            old_manager = match.group(2).strip()
            if old_manager != new_manager:
                html_content = pattern.sub(
                    r'\g<1>' + new_manager + r'\g<3>',
                    html_content,
                    count=1
                )
                updated.append((qr_code, old_manager, new_manager))
            # nếu bằng nhau thì không cần ghi, không báo lỗi
        else:
            skipped.append(qr_code)

    return html_content, updated, skipped


def main():
    print("=" * 60)
    print("  HMO Equipment — Đồng bộ Cán bộ Quản lý")
    print("=" * 60)

    # 1. Đọc xlsx
    print(f"\n📂 Đọc dữ liệu từ: {XLSX_FILE.name}")
    mapping = read_managers_from_xlsx(XLSX_FILE)
    print(f"   → {len(mapping)} thiết bị trong Master_Data")

    # 2. Đọc HTML
    print(f"📄 Đọc file: {HTML_FILE.name}")
    html_content = HTML_FILE.read_text(encoding="utf-8")

    # 3. Backup
    backup_path = backup_html(HTML_FILE)
    print(f"💾 Backup: backups/{backup_path.name}")

    # 4. Cập nhật
    new_html, updated, skipped = update_managers_in_html(html_content, mapping)

    # 5. Ghi file
    HTML_FILE.write_text(new_html, encoding="utf-8")

    # 6. Báo cáo
    print(f"\n✅ Đã cập nhật: {len(updated)} thiết bị")
    if updated:
        print()
        print(f"  {'Mã QR':<22} {'Cũ':<35} {'Mới'}")
        print(f"  {'-'*22} {'-'*35} {'-'*35}")
        for qr, old, new in updated:
            old_display = (old[:33] + '..') if len(old) > 35 else old
            new_display = (new[:33] + '..') if len(new) > 35 else new
            print(f"  {qr:<22} {old_display:<35} {new_display}")

    if skipped:
        print(f"\n⚠️  Không tìm thấy trong HTML ({len(skipped)} thiết bị):")
        for qr in skipped:
            print(f"   - {qr}")

    print(f"\n🎉 Hoàn thành! File đã được cập nhật: {HTML_FILE.name}")
    print("=" * 60)


if __name__ == "__main__":
    main()
